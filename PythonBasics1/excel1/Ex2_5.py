# print("A1 TO C1")
#read by ROW WISE

from openpyxl import load_workbook
from openpyxl import Workbook


# create workbook obj

workbook = load_workbook('response1.xlsx')

#get sheet obj
sheet = workbook.active

print("print multiple selected rows")
rows = sheet[1:3]  # read rows from 1 to 3

print("**********print entire sheet*****************")
for row in rows:
    print("")
    for cell in row:
        print(cell.value , end= " ")






"""
   

print("selected read by row and column")
for row in sheet.iter_rows(min_row=1,
                          max_row=2,
                            min_col=1,
                            max_col=3):
     print(row)

for row in sheet.iter_rows(min_row=1,
                          max_row=2,
                            min_col=1,
                            max_col=3,values_only=True):
 print(row)

"""
