from openpyxl import load_workbook
from openpyxl import Workbook

#Example2
workbook = load_workbook(filename="response.xlsx")

print(workbook.sheetnames)

sheet = workbook.active
print(sheet.title)

 
