"""
1.open existing excel
2.Delete sheet
3.save file


How to create workBook obj for existing excel?
workbook = load_workbook('response.xlsx')


"""

from openpyxl import Workbook, load_workbook

# create workbook obj
workbook = load_workbook('response.xlsx')


# get the existing sheet
#sheet = workbook.get_sheet_by_name("my data")
sheet = workbook ["my data"]

#delete sheet
workbook.remove(sheet)


# save file
workbook.save(filename="response.xlsx")






