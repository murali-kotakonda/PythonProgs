from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('response.xlsx')

sheet = workbook.active

print("**********print entire sheet*****************")
for row in sheet.rows:
    print("")
    for cell in row:
        print(cell.value , end= " ")


 

print("**********print entire sheet*****************")

for row in workbook[ "sheet1" ].rows:
    for cell in row:
        value = cell.value
        print(value)


#or


for row in range( 1, sheet.max_row + 1 ):
    for col in range( 0, sheet.max_column ):
        value = sheet[ row ][ col ].value