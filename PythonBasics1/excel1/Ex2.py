from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('response.xlsx')

sheet = workbook.active


print(sheet["A1"].value)
print(sheet["B1"].value)
print(sheet["C1"].value)


print(sheet["A2"].value)
print(sheet["B2"].value)
print(sheet["C2"].value)


print(sheet.cell(row=1, column=1).value)


print("rows=",sheet.max_row)
print("columns=",sheet.max_column)
